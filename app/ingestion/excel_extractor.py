"""Excel / CSV extraction: every sheet becomes a table; cells flattened to text."""
from __future__ import annotations

from pathlib import Path

from app.core.logging_config import get_logger

log = get_logger(__name__)


def extract_excel(path: Path, doc) -> None:
    if path.suffix.lower() == ".csv":
        _extract_csv(path, doc)
        return
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        log.warning("openpyxl not available: %s", exc)
        return

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        log.warning("Failed to open workbook %s: %s", path.name, exc)
        return

    text_parts: list[str] = []
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                rows.append(cells)
        if rows:
            doc.tables.append(rows)
            text_parts.append(f"# Sheet: {ws.title}")
            text_parts.extend("\t".join(r) for r in rows)
    wb.close()
    doc.pages = ["\n".join(text_parts)]
    doc.full_text = "\n".join(text_parts).strip()


def _extract_csv(path: Path, doc) -> None:
    import csv

    rows: list[list[str]] = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.reader(fh):
                rows.append([c.strip() for c in row])
    except Exception as exc:  # noqa: BLE001
        log.warning("Failed to read CSV %s: %s", path.name, exc)
        return
    if rows:
        doc.tables.append(rows)
    text = "\n".join("\t".join(r) for r in rows)
    doc.pages = [text]
    doc.full_text = text.strip()
